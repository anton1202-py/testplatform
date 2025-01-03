import asyncio
from datetime import datetime
import logging
import tempfile
from openpyxl.styles import Alignment, Border, Side
from django.http import FileResponse, HttpResponse, JsonResponse
from openpyxl import Workbook
import requests
from django.db import transaction
from django.db.models import Count, Prefetch, Q, Case, When, Value, BooleanField, Sum, F
from django.utils import timezone
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.generics import GenericAPIView, ListAPIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.viewsets import ModelViewSet
import telegram

from analyticalplatform.settings import (ADMINS_CHATID_LIST, OZON_ID, TELEGRAM_TOKEN, TOKEN_MY_SKLAD, TOKEN_OZON,
                                         TOKEN_WB, TOKEN_YM)
from core.enums import MarketplaceChoices
from core.models import Account, Platform, User
from core.serializers import ProductsExportSerializer
from unit_economics.integrations import (
    add_marketplace_product_to_db,
    calculate_mp_price_with_incoming_profitability,
    calculate_mp_price_with_profitability,
    calculate_mp_profitability_with_incoming_price, profitability_calculate,
    profitability_calculate_only,
    save_overheds_for_mp_product, calculate_quarantine_mp_products,
    send_message_async, update_price_info_from_user_request)
from unit_economics.models import (MarketplaceAction, MarketplaceCommission,
                                   MarketplaceProduct,
                                   MarketplaceProductInAction,
                                   MarketplaceProductPriceWithProfitability,
                                   ProductPrice,
                                   ProfitabilityMarketplaceProduct, StoreOverhead)
from unit_economics.periodic_tasks import (action_article_price_to_db,
                                           moy_sklad_costprice_add_to_db)
from unit_economics.serializers import (
    AccountSelectSerializer, AccountSerializer, BrandSerializer,
    MarketplaceActionSerializer, MarketplaceCommissionSerializer,
    MarketplaceProductInActionSerializer,
    MarketplaceProductPriceWithProfitabilitySerializer,
    MarketplaceProductSerializer, PlatformSerializer, ProductNameSerializer,
    ProductPriceSelectSerializer, ProductPriceSerializer,
    ProfitabilityMarketplaceProductSerializer, StoreOverheadListSerializer, StoreOverheadSerializer)
from unit_economics.tasks_moy_sklad import (moy_sklad_add_data_to_db, moy_sklad_costprice_calculate_for_bundle,
                                            moy_sklad_stock_data)
from unit_economics.tasks_ozon import (ozon_comission_logistic_add_data_to_db,
                                       ozon_products_data_to_db)
from unit_economics.tasks_wb import (wb_categories_list,
                                     wb_comission_add_to_db,
                                     wb_logistic_add_to_db,
                                     wb_products_data_to_db)
from unit_economics.tasks_yandex import (
    yandex_add_products_data_to_db, yandex_business_list,
    yandex_comission_logistic_add_data_to_db)

logger = logging.getLogger(__name__)


class ProductPriceMSViewSet(viewsets.ViewSet):
    """ViewSet для работы с продуктами на платформе МойСклад"""
    # queryset = ProductPrice.objects.filter(platform=Platform.objects.get(platform_type=MarketplaceChoices.MOY_SKLAD))
    queryset = ProductPrice.objects.all()
    serializer_class = ProductPriceSerializer

    def list(self, request):
        """Получение данных о продуктах из API и обновление базы данных"""
        user = request.user

        account, created = Account.objects.get_or_create(
            user=user,
            platform=Platform.objects.get(
                platform_type=MarketplaceChoices.MOY_SKLAD),
            defaults={
                'name': 'Мой склад',
                'authorization_fields': {'token': TOKEN_MY_SKLAD}
            }
        )
        # Если аккаунт уже существует, но токен не установлен, обновите его
        if not created and account.authorization_fields.get('token') != TOKEN_MY_SKLAD:
            account.authorization_fields['token'] = TOKEN_MY_SKLAD
            account.save()
        total_processed = 0  # Счетчик обработанных записей
        # moy_sklad_add_data_to_db()
        print('moy_sklad_costprice_calculate_for_bundle')
        # action_article_price_to_db()
       
        # yandex_comission_logistic_add_data_to_db()
        moy_sklad_add_data_to_db()
        # wb_products_data_to_db()
        # wb_logistic_add_to_db()
        # wb_comission_add_to_db()
        # ozon_products_data_to_db()
        # yandex_add_products_data_to_db()
        # ozon_comission_logistic_add_data_to_db()
        
        # yandex_comission_logistic_add_data_to_db()
        # moy_sklad_stock_data()
        # profitability_calculate(user_id=user.id)
        # print('moy_sklad_costprice_add_to_db ')
        # moy_sklad_costprice_add_to_db()
        # print('Прошли moy_sklad_costprice_add_to_db ')
        # action_article_price_to_db()
        # profitability_calculate(user.id, overheads=0.2, profitability_group=None, costprice_flag='table')
       
        updated_products = ProductPrice.objects.all()
        serializer = ProductPriceSerializer(updated_products, many=True)
        return Response(
            {'status': 'success', 'message': f'Total processed: {total_processed}',
             'data': serializer.data},
            status=status.HTTP_200_OK)


class ProductMoySkladViewSet(ModelViewSet):
    # queryset = (ProductPrice.objects.filter(platform=Platform.objects.get(platform_type=MarketplaceChoices.MOY_SKLAD))
    #             .annotate(product_count=Count('id')))
    queryset = (ProductPrice.objects.all()
                .annotate(product_count=Count('id')))
    serializer_class = ProductPriceSerializer


class TopSelectorsViewSet(GenericAPIView):
    """
    Получаем информацию по выбору для верхних селекторов на странице

    > Магазин
    > Бренд
    > Товар
    > Маркетплейс
    > Фулфилмент

    На выходе получем словарь с данными:
    {
        accounts: [
            {
                "id": "account_id",
                "name": "account_name",
            }
        ],
        platforms: [
            {
                "id": platform_id,
                "name": "platform_name",
                "platform_type": platform_type
            }
        ],
        brands: [
            {
                "brand": "brand_name",
            }
        ],
        goods:[
            {
                "id": product_id,
                "name": "product_name",
                "brand": "brand_name",
                "vendor": "seller_article"
            }
        ]
    }
    """
    permission_classes = [IsAuthenticated]
    serializer_class = PlatformSerializer

    def get(self, request, *args, **kwargs):
        """
        GET-запрос для получения отфильтрованных данных
        """
        user_id = self.request.query_params.get('user_id')
        accounts_data = Account.objects.filter(
            platform__id__in=[1, 2, 4], user__id=user_id).order_by('id')
        platforms_data = Platform.objects.filter(
            id__in=[1, 2, 4]).order_by('id')
        brands_data = ProductPrice.objects.all().values(
            'brand').distinct().order_by('brand')
        goods_data = ProductPrice.objects.filter(
            account__user__id=user_id).order_by('name')

        top_selection_platform_id = self.request.query_params.get(
            'top_selection_platform_id')
        top_selection_account_id = self.request.query_params.get(
            'top_selection_account_id')
        top_selection_brand = self.request.query_params.get(
            'top_selection_brand')
        top_selection_product_name = self.request.query_params.get(
            'top_selection_product_name')

        # В приоритете верхние фильтры
        if top_selection_platform_id:
            platforms_list = top_selection_platform_id.split(',')
            accounts_data = accounts_data.filter(
                platform__id__in=platforms_list)
            goods_data = goods_data.filter(
                Q(mp_product__platform__id__in=platforms_list)).distinct()
            brands_data = brands_data.filter(
                mp_product__platform__id__in=platforms_list)

        if top_selection_account_id:
            accounts_list = top_selection_account_id.split(',')
            # accounts_data = accounts_data.filter(id__in=accounts_list)
            goods_data = goods_data.filter(
                Q(mp_product__account__id__in=accounts_list)).distinct()
            brands_data = brands_data.filter(
                mp_product__account__id__in=accounts_list)

        if top_selection_brand:
            brands = top_selection_brand.split(',')
            goods_data = goods_data.filter(brand__in=brands)

        if top_selection_product_name:
            products_list = top_selection_product_name.split(',')

        try:

            main_result = {
                "accounts": AccountSelectSerializer(accounts_data, many=True).data,
                "platforms": PlatformSerializer(platforms_data, many=True).data,
                "brands": brands_data,
                "goods": ProductPriceSelectSerializer(goods_data, many=True).data,
            }
            return Response(main_result, status=status.HTTP_200_OK)
        except User.DoesNotExist:
            return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)


class PlatformViewSet(viewsets.ReadOnlyModelViewSet):
    """Получаем платформы(магазины) текущего пользователя"""
    permission_classes = [IsAuthenticated]
    serializer_class = PlatformSerializer

    def get_queryset(self):
        return Platform.objects.filter(accounts__user=self.request.user).distinct()


class AccountViewSet(viewsets.ReadOnlyModelViewSet):
    """Получаем аккаунты текущего пользователя"""
    permission_classes = [IsAuthenticated]
    serializer_class = AccountSerializer

    def get_queryset(self):
        return Account.objects.filter(user=self.request.user)


class BrandViewSet(viewsets.ViewSet):
    """Получаем бренды товаров текущего пользователя"""
    permission_classes = [IsAuthenticated]

    def list(self, request):
        brands = ProductPrice.objects.filter(mp_product__account__user=self.request.user).values_list('brand',
                                                                                                      flat=True).distinct()
        serializer = BrandSerializer(instance=brands, many=True)
        return Response(serializer.data)


class ProductNameViewSet(viewsets.ViewSet):
    """Получаем название товаров текущего пользователя"""
    permission_classes = [IsAuthenticated]

    def list(self, request):
        product_names = (ProductPrice.objects.filter(mp_product__account__user=self.request.user)
                         .values_list('name', flat=True).distinct())
        serializer = ProductNameSerializer(instance=product_names, many=True)
        return Response(serializer.data)


class MarketplaceProductViewSet(viewsets.GenericViewSet):
    """Получаем товары для выбранной платформы + фильтрация по данным от пользователя
       + поля для поиска 'name', 'barcode' пример запроса GET /api/marketplace-products/?search=123456789
       + поля для сортировки 'profit', 'profitability' пример запроса
       GET /api/marketplace-products/?ordering=mp_profitability__profit
       (или -profit для сортировки по убыванию)
    """
    permission_classes = [IsAuthenticated]
    serializer_class = MarketplaceProductSerializer
    # Подключаем поиск и сортировку
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['name', 'barcode']  # Поля для поиска
    # Поля для сортировки ['profit', 'profitability']
    ordering_fields = ['mp_profitability__profit',
                       'mp_profitability__profitability']

    def get_queryset(self):
        user = self.request.user
        queryset = MarketplaceProduct.objects.filter(
            account__user=user, is_active=True).select_related('mp_product_profit_price')
        top_selection_platform_id = self.request.query_params.get(
            'top_selection_platform_id')
        top_selection_account_id = self.request.query_params.get(
            'top_selection_account_id')
        top_selection_brand = self.request.query_params.get(
            'top_selection_brand')
        top_selection_product_name = self.request.query_params.get(
            'top_selection_product_name')
        quarantine = self.request.query_params.get(
            'quarantine')
        self.quarantine_percent = self.request.query_params.get(
            'quarantine_percent')
        # Повторяющиеся фильтры в верху страницы и вверху таблицы.
        table_platform_id = self.request.query_params.get('table_platform_id')
        table_brand = self.request.query_params.get('table_brand')
        
        filter_platform_id = ''
        filter_brands = ''
        # В приоритете верхние фильтры
        if top_selection_platform_id:
            filter_platform_id = top_selection_platform_id
        elif table_platform_id:
            filter_platform_id = table_platform_id
        

        if filter_platform_id:
            platforms_list = filter_platform_id.split(',')
            queryset = queryset.filter(platform__id__in=platforms_list)
        if top_selection_account_id:
            accounts_list = top_selection_account_id.split(',')
            queryset = queryset.filter(account__id__in=accounts_list)
        if top_selection_brand:
            brands = top_selection_brand.split(',')
            queryset = queryset.filter(product__brand__in=brands)
        if top_selection_product_name:
            products_list = top_selection_product_name.split(',')
            queryset = queryset.filter(product__in=products_list)
        self.all_product_count = len(queryset)
        if quarantine:
            percent = 20
            if self.quarantine_percent:
                percent = int(self.quarantine_percent)

            percent = percent / 100
            great = 1 + percent
            less = 1 - percent
            queryset = queryset.annotate(
                wb_price=F('product__price_product__wb_price'),
                yandex_price=F('product__price_product__yandex_price'),
                rrc=F('product__price_product__rrc')
            )
            
            if '1' in platforms_list:
                queryset = queryset.filter(
                    Q(wb_price__gt=F('rrc') * great) | Q( wb_price__lt=F('rrc') * less)
                )
            elif '2' in platforms_list:
                queryset = queryset.filter(
                    Q(yandex_price__gt=F('rrc') * great) | Q(yandex_price__lt=F('rrc') * less)
                )
            elif '4' in platforms_list:
                queryset = queryset.filter(
                    Q(product__price_product__rrc__gt=F('product__ozon_price_product__ozon_price') * great) | Q(product__price_product__rrc__lt=F('product__ozon_price_product__ozon_price')* less)
                )
        # Добавляем prefetch_related для акции
        queryset = queryset.prefetch_related(
            Prefetch('product_in_action',
                     queryset=MarketplaceProductInAction.objects.select_related('action'))
        )
        return queryset.order_by('id')

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        profitability_group = request.query_params.get('profitability_group')
        calculate_product_price = request.query_params.get(
            'calculate_product_price')
        action_id = request.query_params.get('action_id')
        costprice_flag = request.query_params.get('costprice_flag')
        price_toggle = request.query_params.get('price_toggle')
        order_delivery_type = request.query_params.get('order_delivery_type')

        if profitability_group:
            # Срабатывает, когда нажимают на бар диаграммы, чтобы отфильтровать по нему товары
            result = profitability_calculate(
                request.user.id, profitability_group=profitability_group, costprice_flag=costprice_flag, order_delivery_type=order_delivery_type)
            queryset = queryset.filter(
                id__in=[p.id for p in result['filtered_products']])
        
        if price_toggle:
            # Срабатывает, когда переключатель ЦЕНА в положении МОЙ СКЛАД
            result = profitability_calculate_only(
                queryset, costprice_flag=costprice_flag, order_delivery_type=order_delivery_type)
            queryset = MarketplaceProduct.objects.filter(
                id__in=[p.id for p in result])
           
        # Фильтр для Пересчета цены на основании входящей рентабельности. Сохраняет цену и рентабельность
        if calculate_product_price:
            # Срабатывает, когда переключатель ЦЕНА в положении ПО УРОВНЮ РЕНТАБЕЛЬНОСТИ
            updated_products = calculate_mp_price_with_incoming_profitability(
                float(calculate_product_price), queryset, costprice_flag=costprice_flag, order_delivery_type=order_delivery_type)
            queryset = MarketplaceProduct.objects.filter(
                id__in=[p.id for p in updated_products])

        # Фильтр по id акции
        if action_id:
            # Срабатывает, когда переключатель нужно получить товары в АКЦИИ с номер action_id
            queryset = queryset.filter(
                product_in_action__action__id=action_id).distinct()
            # Аннотируем каждый товар значением status из связанной модели MarketplaceProductInAction
            queryset = queryset.annotate(
                has_participation=Case(
                    When(product_in_action__status=True, then=Value(True)),
                    default=Value(False),
                    output_field=BooleanField()
                )
            ).order_by('-has_participation')

            # Пересчитывает рентабельность и прибыль на основании входящей цены. И сохраняет цену и рентабельность
            updated_profitability = calculate_mp_profitability_with_incoming_price(action_id,
                queryset, costprice_flag=costprice_flag, order_delivery_type=order_delivery_type)
            queryset = MarketplaceProduct.objects.filter(
                id__in=[p.id for p in updated_profitability])

        # Получение параметра сортировки из запроса
        ordering = request.query_params.get('ordering', None)
        if ordering:
            queryset = queryset.order_by(ordering)
        else:
            queryset = queryset.order_by('id')
        self.excel_serializer = self.get_serializer(queryset, many=True)
        page = self.paginate_queryset(queryset)
        serializer = self.get_serializer(queryset, many=True)
        response_data = {
            'quarantine_count': len(calculate_quarantine_mp_products(self.quarantine_percent, queryset)),
            'all_products_count': self.all_product_count,
            'data_count': len(serializer.data),
            'data': serializer.data,
        }

        return Response(response_data)

        # print('page', page)
        # if page is not None:
        #     serializer = self.get_serializer(page, many=True)
        #     print(serializer.data)
        #     return self.get_paginated_response(serializer.data)

        # serializer = self.get_serializer(queryset, many=True)
        # return Response(serializer.data)
    
    def post(self, request, *args, **kwargs):
        
        self.list(request)
        serialized_data = self.excel_serializer.data

        price_toggle = self.request.query_params.get('price_toggle', '')
        top_selection_platform_id = self.request.query_params.get(
            'top_selection_platform_id')
        table_platform_id = self.request.query_params.get('table_platform_id')
        costprice_flag = self.request.query_params.get('costprice_flag', '')
        ff_type = self.request.query_params.get('order_delivery_type', '')
        account_ids = self.request.query_params.get(
            'top_selection_account_id')
        platform_ids = ''

        # В приоритете верхние фильтры
        if top_selection_platform_id:
            platform_ids = top_selection_platform_id.split(',')
        elif table_platform_id:
            platform_ids = table_platform_id.split(',')

        if price_toggle:
            price_toggle = price_toggle
        if costprice_flag:
             costprice_flag = costprice_flag
        platform_name = ''
        if platform_ids:
            for platform_id in platform_ids:
                platform_name += f'{Platform.objects.get(id=int(platform_id)).name} '
        overheads = ''
        account_name = ''
        if account_ids:
            account_ids = account_ids.split(',')
            if len(account_ids) == 1:
                overheads_data = StoreOverhead.objects.filter(account__id=account_ids[0]).aggregate(
                overhead_sum=Sum('overhead'))
                overheads = overheads_data['overhead_sum']
                account_name = Account.objects.get(id=int(account_ids[0])).name
            else:
                for account_id in account_ids:
                    account_name += f'{Account.objects.get(id=int(account_id)).name} '
                
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value=f'Marketpalce: {platform_name}')
        ws.cell(row=2, column=1, value=f'Магазин: {account_name}')
        ws.cell(row=3, column=1, value=f'Накладные расходы: {overheads}')
        ws.cell(row=4, column=1, value=f'Тип ФФ: {ff_type}')
    
        table_headers = (
            "Название на маркетплейсе",
            "Название на 'Мой склад'",
            "Артикул на маркетплейсе",
            "Артикул на 'Мой склад'",
            "SKU на маркетплейсе",
            "SKU на 'Мой склад'",

            "Баркод",
            "Бренд",
            "РРЦ",
            "Цена",

            "Себестоимость",
            "Комиссия",
            "Логистика",
            "Прибыль/убыток",
            "Рентабельность",
        )
        
        for col, header in enumerate(table_headers, start=1):
            ws.cell(row=5, column=col, value=header)
    
        for row, data in enumerate(serialized_data, start=6):
            marketplace_prod_obj = MarketplaceProduct.objects.get(id=data['id'])
            commission = 0
            platform_id = marketplace_prod_obj.platform.id
            if not price_toggle and costprice_flag == 'table':
                price = data['usual_price']
            elif not price_toggle and costprice_flag == 'enter':
                price = data['profit_price']
            else:
                price = data['price']
            # print(data)
            if data['commission']:
                commission = data['commission']['fbs_commission']
                if ff_type == 'fbs':
                    commission = data['commission']['fbs_commission']
                elif ff_type == 'fbo':
                    commission = data['commission']['fbo_commission']
                elif ff_type == 'dbs':
                    commission = data['commission']['dbs_commission']
                elif ff_type == 'fbs_express':
                    commission = data['commission']['fbs_express_commission']
                
            if data['logistic_cost']:
                if platform_id == 4 and ff_type == 'fbs':
                    logistic_cost = data['logistic_cost']['cost_logistic_fbs']
                elif platform_id == 4 and ff_type == 'fbo':
                    logistic_cost = data['logistic_cost']['cost_logistic_fbo']
                else:
                    logistic_cost = data['logistic_cost']['cost_logistic']
            else:
                logistic_cost = 0
            if costprice_flag == 'enter':
                costprice = data['posting_costprice']
            else:
                costprice = data['cost_price']

            ws.cell(row=row, column=1, value=marketplace_prod_obj.name)
            ws.cell(row=row, column=2, value=marketplace_prod_obj.product.name)
            ws.cell(row=row, column=3, value=marketplace_prod_obj.seller_article)
            ws.cell(row=row, column=4, value=marketplace_prod_obj.product.vendor)
            ws.cell(row=row, column=5, value=marketplace_prod_obj.sku)
            ws.cell(row=row, column=6, value=marketplace_prod_obj.product.code)

            ws.cell(row=row, column=7, value=data['barcode'])
            ws.cell(row=row, column=8, value=data['brand'])
            ws.cell(row=row, column=9, value=data['rrc'])
            ws.cell(row=row, column=10, value=price)

            ws.cell(row=row, column=11, value=costprice)
            ws.cell(row=row, column=12, value=(commission*price)/100)
            ws.cell(row=row, column=13, value=logistic_cost)
            ws.cell(row=row, column=14, value=data['profit'])
            ws.cell(row=row, column=15, value=data['profitability'])
        
        al = Alignment(horizontal="center",
                   vertical="center")
        thin = Side(border_style="thin", color="000000")

        ws.column_dimensions['A'].width = 23
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18

        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['I'].width = 18
        ws.column_dimensions['J'].width = 18

        ws.column_dimensions['K'].width = 18
        ws.column_dimensions['L'].width = 18
        ws.column_dimensions['M'].width = 18
        ws.column_dimensions['N'].width = 18
        ws.column_dimensions['O'].width = 18

        for i in range(len(serialized_data)+1):
            for c in ws[f'A{i+5}:O{i+5}']:
                for i in range(15):
                    c[i].border = Border(top=thin, left=thin,
                                         bottom=thin, right=thin)
                    c[i].alignment = al
    
        temp_file = tempfile.mktemp(prefix=f"Сводная таблица {datetime.now().strftime('%Y-%m-%d ')}", suffix=".xlsx")
        try:
            wb.save(temp_file)
        except Exception:
            return FileResponse()
        return FileResponse(open(temp_file, "rb"), as_attachment=True)


class ProfitabilityAPIView(GenericAPIView):
    """
    API для расчета рентабельности и сохранения накладных расходов.
    """
    serializer_class = ProfitabilityMarketplaceProductSerializer

    def get(self, request, *args, **kwargs):
        """
        GET-запрос для расчета рентабельности всех товаров пользователя.
        """
        user_id = self.kwargs.get('user_id')

        queryset = ProfitabilityMarketplaceProduct.objects.filter(
            Q(mp_product__account__user__id=user_id) & Q(mp_product__is_active=True))
    
        product_situations = ProductPrice.objects.filter(
            Q(account__user__id=user_id) & Q(mp_product__is_active=True))
        
        top_selection_platform_id = self.request.query_params.get(
            'top_selection_platform_id')
        top_selection_account_id = self.request.query_params.get(
            'top_selection_account_id')
        top_selection_brand = self.request.query_params.get(
            'top_selection_brand')
        top_selection_product_name = self.request.query_params.get(
            'top_selection_product_name')

        # В приоритете верхние фильтры
        if top_selection_platform_id:
            platforms_list = top_selection_platform_id.split(',')
            queryset = queryset.filter(
                Q(mp_product__platform__id__in=platforms_list))
            product_situations = product_situations.filter(
                Q(mp_product__platform__id__in=platforms_list) & Q(mp_product__is_active=True)).distinct()
        if top_selection_account_id:
            accounts_list = top_selection_account_id.split(',')
            queryset = queryset.filter(
                mp_product__account__id__in=accounts_list)
            product_situations = product_situations.filter(
                Q(mp_product__account__id__in=accounts_list) & Q(mp_product__is_active=True))

        if top_selection_brand:
            brands = top_selection_brand.split(',')
            queryset = queryset.filter(mp_product__product__brand__in=brands)
            product_situations = product_situations.filter(brand__in=brands)

        if top_selection_product_name:
            products_list = top_selection_product_name.split(',')
            queryset = queryset.filter(
                mp_product__product__id__in=products_list)
            product_situations = product_situations.filter(
                id__in=products_list)

        try:
            print('len(queryset)', len(queryset))
            result = queryset.aggregate(
                count_above_20=Count('id', filter=Q(profitability__gt=20)),
                count_between_10_and_20=Count('id', filter=Q(
                    profitability__lte=20) & Q(profitability__gt=10)),
                count_between_0_and_10=Count('id', filter=Q(
                    profitability__lte=10) & Q(profitability__gt=0)),
                count_between_0_and_minus_10=Count('id', filter=Q(
                    profitability__lte=0) & Q(profitability__gt=-10)),
                count_between_minus_10_and_minus_20=Count('id', filter=Q(
                    profitability__lte=-10) & Q(profitability__gt=-20)),
                count_below_minus_20=Count(
                    'id', filter=Q(profitability__lte=-20)),
            )
            all_situations = len(queryset)
            product_situations = len(product_situations)
            minus_situations = result['count_between_0_and_minus_10'] + result['count_between_minus_10_and_minus_20'] + result['count_below_minus_20']
            main_result = {
                'diagram_data': result,
                'minus_situations': minus_situations,
                'all_situations': all_situations,
                'product_situations': product_situations

            }
            return Response(main_result, status=status.HTTP_200_OK)
        except User.DoesNotExist:
            return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)

    def post(self, request, *args, **kwargs):
        """
        POST-запрос для обновления накладных расходов и пересчета рентабельности.
        Входящие данные:
        mp_product_dict: словарь типа {mp_product_id: product_overheads}
        mp_product_id - id продукта из таблицы MarketplaceProduct
        product_overheads - накладные расходы в формате float (например 0.2)
        """
        overheads_data = request.data.get('overheads_data', {})
        user_id = request.data.get('user_id')

        if not overheads_data or not user_id:
            return Response({"error": "Invalid data"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Сохраняем накладные расходы
            save_overheds_for_mp_product(overheads_data)

            # Пересчитываем рентабельность
            result = profitability_calculate(user_id)

            return Response(result, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class UpdatePriceView(GenericAPIView):
    """
    Представление для обновления цен на товары и накладных расходов.
    Обновляет цены на Мой склад и в БД, если пользователь отправил запрос
    {
        'user_id': user_id,
        'quarantine_percent': quarantine_percent,
        'products_data': [
            {
                'marketplaceproduct_id': marketplaceproduct_id,
                'new_price': new_price
            }
        ]
    }
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        try:
            # Вызов вашей функции для обновления цен
            update_price_info_from_user_request(request.data)
            return Response({"detail": "Цены успешно обновлены"}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class CalculateMarketplacePriceView(GenericAPIView):
    """
    Вычисление цены на маркетплейсе на основе рентабельности
    в body нужно прокинуть user_id пользователя.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        user_id = request.user.id
        calculate_mp_price_with_profitability(user_id)
        return Response({"detail": "Цены успешно обновлены."})


class MarketplaceActionListView(APIView):
    """Получение всех акций и товаров для указанной платформы"""
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        # Получаем platform_id из параметров запроса
        platform_id = request.query_params.get('platform_id')
        action_id = request.query_params.get('action_id')
        if not platform_id:
            platform_id = 1
        try:
            platform = Platform.objects.get(id=platform_id)
        except Platform.DoesNotExist:
            return Response({"detail": "Платформа не найдена"}, status=status.HTTP_404_NOT_FOUND)
        # Получаем акции для выбранной платформы
        if action_id:
            actions = MarketplaceAction.objects.filter(
                platform=platform, id=action_id).prefetch_related('action')
        else:
            actions = MarketplaceAction.objects.filter(
                platform=platform).prefetch_related('action')
        platform_data = {
            'platform_id': platform.id,
            'platform_name': platform.name,
            'actions': []
        }
        for action in actions:
            # Получаем продукты для каждой акции
            products_in_action = MarketplaceProductInAction.objects.filter(action=action).select_related(
                'marketplace_product')
            # Пропускаем акцию, если у нее нет продуктов
            if not products_in_action.exists():
                continue
            products_data = MarketplaceProductInActionSerializer(
                products_in_action, many=True).data
            action_data = {
                'action_id': action.id,
                'action_number': action.action_number,
                'action_name': action.action_name,
                'date_start': action.date_start,
                'date_finish': action.date_finish,
                'products': products_data,
            }
            platform_data['actions'].append(action_data)
        return Response(platform_data, status=status.HTTP_200_OK)


class CalculateMPPriceView(APIView):
    """
    Эндпоинт для расчета цены товара на маркетплейсе на основе рентабельности.

    Входящие данные (в теле POST-запроса):
        incoming_profitability: float - входящая рентабельность с которой сравниваем рентабельность из БД
        product_ids: list - список ID товаров, которые находятся на странице

    Возвращает:
        Список объектов модели MarketplaceProduct с обновленными ценами.
    """

    def post(self, request, *args, **kwargs):
        incoming_profitability = request.data.get('incoming_profitability')
        product_ids = request.data.get('product_ids')

        if not incoming_profitability or not product_ids:
            return Response({"detail": "Нужно указать входящую рентабельность и список ID товаров"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            product_ids = [int(pid) for pid in product_ids]
        except ValueError:
            return Response({"detail": "Не верный формат ID товаров"}, status=status.HTTP_400_BAD_REQUEST)

        mp_products_list = calculate_mp_price_with_incoming_profitability(
            incoming_profitability, product_ids)

        serializer = MarketplaceProductSerializer(mp_products_list, many=True)
        return Response(serializer.data)


class MarketplaceProductPriceWithProfitabilityViewSet(viewsets.ReadOnlyModelViewSet):
    """    
    profit_price - это по Fifo
    usual_price = простая рентабельность
    """
    permission_classes = [IsAuthenticated]
    serializer_class = MarketplaceProductPriceWithProfitabilitySerializer
    filter_backends = [SearchFilter]
    # Поле для фильтрации по бренду
    search_fields = ['mp_product__product__brand']

    def get_queryset(self):
        queryset = MarketplaceProductPriceWithProfitability.objects.all()

        # Фильтрация по бренду, если передан параметр 'brand'
        brand = self.request.query_params.get('brand')
        if brand:
            queryset = queryset.filter(mp_product__product__brand=brand)

        return queryset


class UserIdView(APIView):
    """Апишка, которая отдаёт фронту id юзера"""
    permission_classes = [IsAuthenticated]

    def get(self, request, format=None):
        user_id = request.user.id
        return Response({'user_id': user_id})


class MarketplaceActionList(ListAPIView):
    """
    Выводит список акций, в которых участвуют наши товары.
    """
    permission_classes = [IsAuthenticated]
    serializer_class = MarketplaceActionSerializer

    def get_queryset(self):
        today = timezone.now().date()
        # Получаем параметры из запроса
        user_id = self.request.query_params.get('user_id')
        account_id = self.request.query_params.get('account_id')
        platform_id = self.request.query_params.get('platform_id')
        # Фильтруем акции, которые ещё не закончились
        queryset = MarketplaceAction.objects.filter(
            id__in=MarketplaceProductInAction.objects.values_list('action__id', flat=True).distinct(),
            date_finish__gte=today)
        if user_id:
            queryset = queryset.filter(account__user_id=user_id)
        if account_id:
            queryset = queryset.filter(account__id=account_id)
        if platform_id:
            queryset = queryset.filter(platform__id=platform_id)
        return queryset


class UpdateMarketplaceProductFlag(APIView):
    """
    Представление для обновления флага change_price_flag на false для списка товаров
    """
    def post(self, request, *args, **kwargs):
        product_ids = request.data.get('product_ids', [])
        if not product_ids:
            return Response({"detail": "Не передан список товаров"}, status=status.HTTP_400_BAD_REQUEST)

        updated_count = MarketplaceProduct.objects.filter(id__in=product_ids).update(change_price_flag=False)

        return Response({"detail": f"Флаг change_price_flag обновлен на false для {updated_count} товара/ов"},
                        status=status.HTTP_200_OK)


class StoreOverheadViewSet(viewsets.ModelViewSet):
    serializer_class = StoreOverheadSerializer

    def get_queryset(self):
        account_id = self.kwargs.get('account_id')
        return StoreOverhead.objects.filter(account_id=account_id)

    def create(self, request, *args, **kwargs):
        account_id = self.kwargs.get('account_id')
        request.data['account'] = account_id
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        account_id = self.kwargs.get('account_id')
        request.data['account'] = account_id
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        account_id = self.kwargs.get('account_id')
        request.data['account'] = account_id
        return super().partial_update(request, *args, **kwargs)

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        total_overhead = sum(overhead.overhead for overhead in queryset if overhead.overhead is not None)
        response_data = {
            'store_overheads': queryset,
            'total_overhead': total_overhead
        }
        return Response(StoreOverheadListSerializer(response_data).data)
