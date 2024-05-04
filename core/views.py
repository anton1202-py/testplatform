import tempfile

from django.http import FileResponse
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl.workbook import Workbook
from rest_framework import filters, mixins, status, views, viewsets
from rest_framework.response import Response

from core.enums import MarketplaceChoices
from core.models import Account, Platform, Product
from core.query_selectors import get_user_accounts
from core.serializers import (
    AccountCreateSerializer,
    AccountsListSerializers,
    ProductManualConnectionCreationSerializer,
    ProductsExportSerializer,
    ProductsListSerializers,
)
from core.utils import sort_products_key


class AccountsViewSet(
    viewsets.GenericViewSet,
    mixins.ListModelMixin,
):
    http_method_names = ["get"]
    filter_backends = [
        filters.OrderingFilter,
        DjangoFilterBackend,
        filters.SearchFilter,
    ]
    serializer_class = AccountsListSerializers

    def get_queryset(self):
        accounts = get_user_accounts(user=self.request.user)

        return accounts


class ProductViewSet(viewsets.GenericViewSet, mixins.ListModelMixin):
    http_method_names = ["get"]
    filter_backends = [filters.OrderingFilter, DjangoFilterBackend, filters.SearchFilter]
    serializer_class = ProductsListSerializers
    filterset_fields = {
        "account": ["exact", "in"],
        "account__user__email": ["exact"],
        "account__platform__platform_type": ["exact", "in"],
        "connection": ["isnull",],
        "name": ["icontains"],
    }

    search_fields = ["name"]

    ordering_fields = ["name"]

    def get_queryset(self):
        queryset = Product.objects.filter(account__user=self.request.user).select_related("connection")

        if self.request.GET.get("only-no-connections"):
            queryset = queryset.filter(has_manual_connection=False).select_related("account__platform")
            exclude_barcodes = []
            for item in queryset:
                required_platform_types = (
                    [MarketplaceChoices.MOY_SKLAD]
                    if item.account.platform.platform_type != MarketplaceChoices.MOY_SKLAD
                    else MarketplaceChoices.values[:-1]
                )
                if item.connection and item.connection.account.platform.platform_type in required_platform_types:
                    exclude_barcodes.append(item.barcode)
            exclude_barcodes = list(set(exclude_barcodes))
            queryset = queryset.exclude(barcode__in=exclude_barcodes)
        return queryset

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        page = self.paginate_queryset(queryset)
        page_ids = [item.id for item in page]
        page_qs = Product.objects.filter(id__in=page_ids)
        sort_by = request.query_params.get("sort_by", "market")

        if page is not None:
            serializer = self.get_serializer(page, many=True)
            serializer_data = serializer.data
            if request.GET.get("no-comparsion"):
                return super().list(request, *args, **kwargs)
            result = []
            analogues = []
            for idx, item in enumerate(page_qs.exclude(account__platform__platform_type=MarketplaceChoices.MOY_SKLAD)):
                analogue = item.connection

                if analogue:
                    analogues.append(analogue.id)

                result.append(
                    {
                        "other_marketplace": serializer_data[idx],
                        "moy_sklad": self.serializer_class(analogue, many=False).data if analogue else None,
                    }
                )

            for not_listed_product in page_qs.filter(
                account__platform__platform_type=MarketplaceChoices.MOY_SKLAD,
            ).exclude(id__in=analogues):
                result.append(
                    {
                        "other_marketplace": None,
                        "moy_sklad": self.serializer_class(not_listed_product, many=False).data,
                    }
                )

            result.sort(
                key=lambda product: sort_products_key(sort_by, product),
                reverse="-" not in sort_by,
            )

            return self.get_paginated_response(result)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class GetMarketplaceTypesAPIView(views.APIView):
    def get(self, request):
        if request.GET.get("with_moy_sklad"):
            data = MarketplaceChoices.labels
        else:
            data = MarketplaceChoices.labels[:-1]
        return Response(data=data)


class CreateAccountAPIView(views.APIView):
    serializer_class = AccountCreateSerializer

    def post(self, request):
        serializer = self.serializer_class(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        platform = Platform.objects.filter(platform_type=serializer.validated_data.pop("platform_type")).first()

        if not platform:
            return Response(status=status.HTTP_400_BAD_REQUEST)

        Account.objects.create(user=request.user, platform=platform, **serializer.validated_data)
        return Response(serializer.validated_data, status=status.HTTP_201_CREATED)


class ProductManualConnectionCreationAPIView(views.APIView):
    serializer_class = ProductManualConnectionCreationSerializer

    def post(self, request):
        serializer = self.serializer_class(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        other_marketplace_product = Product.objects.filter(
            id=serializer.validated_data["other_marketplace_product"]
        ).first()
        moy_sklad_product = Product.objects.filter(id=serializer.validated_data["moy_sklad_product"]).first()

        if not other_marketplace_product or not moy_sklad_product:
            return Response(status=status.HTTP_400_BAD_REQUEST)

        if (
            moy_sklad_product.account.platform.platform_type != MarketplaceChoices.MOY_SKLAD
            or other_marketplace_product.account.platform.platform_type == MarketplaceChoices.MOY_SKLAD
        ):
            return Response(status=status.HTTP_400_BAD_REQUEST)

        other_marketplace_product.connection = moy_sklad_product
        moy_sklad_product.has_manual_connection = True
        other_marketplace_product.has_manual_connection = True
        moy_sklad_product.save()
        other_marketplace_product.save()
        return Response(status=status.HTTP_200_OK)


class GetPlatformAuthFieldsDescriptionAPIView(views.APIView):
    def get(self, request, *args, **kwargs):
        platform_type = self.kwargs.get("platform_type")
        if platform_type is None:
            return Response(status=status.HTTP_404_NOT_FOUND)

        platform = Platform.objects.filter(platform_type=platform_type).first()

        if not platform:
            return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(platform.auth_fields_description)


class ExportReportAPIView(views.APIView):
    def post(self, request, *args, **kwargs):
        serializer = ProductsExportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        wb = Workbook()
        ws = wb.active

        table_headers = (
            "Название на маркетплейсе",
            "Название на 'Мой склад'",
            "Артикул на маркетплейсе",
            "Артикул на 'Мой склад'",
            "SKU на маркетплейсе",
            "SKU на 'Мой склад'",
            "Баркод",
        )
        for col, header in enumerate(table_headers, start=1):
            ws.cell(row=1, column=col, value=header)

        products = Product.objects.filter(id__in=data["products"])
        marketplace_products = products.exclude(account__platform__platform_type=MarketplaceChoices.MOY_SKLAD)
        index = marketplace_products.count()
        analogues = []

        for idx, item in enumerate(marketplace_products):
            analogue = Product.objects.filter(
                account__platform__platform_type=MarketplaceChoices.MOY_SKLAD,
                barcode=item.barcode,
            ).first()

            if analogue:
                analogues.append(analogue.id)
                item_data = [
                    item.name,
                    analogue.name,
                    item.vendor,
                    analogue.vendor,
                    item.sku,
                    analogue.sku,
                    item.barcode,
                ]
            else:
                item_data = [item.name, "", item.vendor, "", item.sku, "", item.barcode]

            for col, product_data in enumerate(item_data, start=1):
                ws.cell(row=2 + idx, column=col, value=product_data)

        for not_listed_product in products.filter(
            account__platform__platform_type=MarketplaceChoices.MOY_SKLAD,
        ).exclude(id__in=analogues):
            item_data = [
                "",
                not_listed_product.name,
                "",
                not_listed_product.vendor,
                "",
                not_listed_product.sku,
                not_listed_product.barcode,
            ]

            for col, product_data in enumerate(item_data, start=1):
                ws.cell(row=2 + index, column=col, value=product_data)

        temp_file = tempfile.mktemp(suffix=".xlsx")

        try:
            wb.save(temp_file)
        except Exception:
            return FileResponse()

        return FileResponse(open(temp_file, "rb"), as_attachment=True)
